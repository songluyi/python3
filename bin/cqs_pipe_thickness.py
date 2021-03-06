# -*- coding: utf-8 -*-
# 2016/7/20 13:37
"""
-------------------------------------------------------------------------------
Function:   插入管道厚度到数据库里
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from cqs_pipe_thickness_database import *
from cqs_pt_rating_database import compliment
from cqs_pt_rating import cqs_pt_rating#这里主要为了引入get_path函数
from cqs_branch_connect_database import return_domain_username
import time
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
header_name=['PIPE_ID', 'BATCH_ID', 'PIPE_ORDER_NUMBER', 'PIPING_MATL_CLASS', 'PIPE_DN', 'PIPE_OUTER',
             'PIPE_THICKNESS', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY', 'LAST_UPDATE_DATE',
             'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2', 'ATTRIBUTE3',
             'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15']

class cqs_pipe_thickness(object):
    def make_exceldata(self,name_list,bug_pipe_id,pipe_id,batch_id,pipe_order_number,domain_name):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'pipe_thickness'
        line=1#从第二行开始
        pipe_id=pipe_id+1
        batch_id=batch_id+1
        for i in range(1,29):
                ws_write.cell(row=1, column=i).value = header_name[i-1]
        for name in name_list:
            # load_name=os.path.basename(name)
            if bug_pipe_id>pipe_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_pi_id来解决下一个文件的pi_id的取值问题
                pipe_id=bug_pipe_id
            wb_load = load_workbook(filename=name)
            sheets = wb_load.get_sheet_names()
            del sheets[-1]#用openpyxl这个模块获取excel标签尾部会多一个页签，具体原因不明
            for page_name in sheets[1::2]:#循环样例应该为16.2 35.2 等
                ws_load = wb_load.get_sheet_by_name(page_name)
                make_column=[x for x in range(4,35)][::2]
                catalog_line=[]
                for search in range(1,100):#动态确定DN位置并且方便循环
                    if "DN" in str(ws_load.cell(row=search, column=1).value):
                        catalog_line.append(search)
                #[5,8,11]
                for row in catalog_line:
                    for column in make_column:
                        if ws_load.cell(row=row, column=column).value is not None:
                            pipe_id+=1
                            pipe_order_number+=1
                            line+=1
                            bug_pipe_id+=1
                            ws_write.cell(row=line, column=1).value=pipe_id
                            ws_write.cell(row=line, column=2).value=batch_id
                            ws_write.cell(row=line, column=3).value=pipe_order_number
                            ws_write.cell(row=line, column=4).value=ws_load.cell(row=1, column=26).value#写入metal class
                            # if ws_load.cell(row=row, column=column).value is None:
                            #     ws_write.cell(row=line, column=5).value='9999'
                            # else:
                            ws_write.cell(row=line, column=5).value=ws_load.cell(row=row, column=column).value#写入PIPE_DN
                            ws_write.cell(row=line, column=6).value=ws_load.cell(row=row+1, column=column).value#PIPE__OUT
                            ws_write.cell(row=line, column=7).value=ws_load.cell(row=row+2, column=column).value#PIPE_THICKNESS
                            ws_write.cell(row=line, column=8).value=domain_name#写入created by
                            # ws_write.cell(row=line, column=9).number_format='yyyy-mm-dd'
                            ws_write.cell(row=line, column=9).value=today_time
                            ws_write.cell(row=line, column=10).value=0#写入last_update_by
                            # ws_write.cell(row=line, column=11).number_format='yyyy-mm-dd'
                            ws_write.cell(row=line, column=11).value=today_time
                            ws_write.cell(row=line, column=12).value=0#写入last_update_login
        name='new'+'管道材料等级表-外径壁厚表'+'.xlsx'
        wb_write.save(name)
        print('已经完成管道材料等级表-外径壁厚表的excel生成')
if __name__ == '__main__':
    cqs=cqs_pipe_thickness()
    name_list=cqs_pt_rating().get_path()
    pipe_id=get_pipeid()
    if pipe_id is None:
        pipe_id=0
    batch_id=get_batch_id()
    if batch_id is None:
        batch_id=0
    pipe_order_number=get_order_number()
    if pipe_order_number is None:
        pipe_order_number=0
    bug_pipe_id=0
    domain_name=return_domain_username()
    cqs.make_exceldata(name_list,bug_pipe_id,pipe_id,batch_id,pipe_order_number,domain_name)
    excel_name='new管道材料等级表-外径壁厚表.xlsx'
    data_list=compliment(header_name,excel_name)
    start_time=time.time()
    insert_db(data_list)
    end_time=time.time()
    print('耗时为：',end_time-start_time,'插入总数为：',len(data_list))
    print('已经完成对管道壁厚表的插入，谢谢使用')
    print('如果未出现完成插入的提示则说明该表导入失败')
    time.sleep(5)



